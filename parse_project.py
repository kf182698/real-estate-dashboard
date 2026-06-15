#!/usr/bin/env python3
"""
單案週檢討解析器 — 由 YAML 對映設定檔驅動
支援多版型偵測、冪等合併到 db.json
"""
import re, json, pathlib, datetime
import openpyxl

try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False

SCRIPT_DIR   = pathlib.Path(__file__).parent
MAPPING_DIR  = SCRIPT_DIR / "parsers"
DB_JSON      = SCRIPT_DIR / "db.json"
WEEK_PAT     = re.compile(r"^\d{4}-\d{4}$")

# ── YAML 載入（若無 pyyaml 則用 json fallback）─────────────────
def load_mapping(project_name: str) -> dict:
    path = MAPPING_DIR / f"{project_name}_mapping.yaml"
    if not path.exists():
        raise FileNotFoundError(f"找不到對映設定檔：{path}")
    if HAS_YAML:
        import yaml
        return yaml.safe_load(path.read_text(encoding="utf-8"))
    # pyyaml 未安裝 → 用簡易解析（只讀 cell 座標）
    raise ImportError("請先 pip install pyyaml")

# ── 工具函式 ──────────────────────────────────────────────────
def _num(v, typ="float"):
    if v is None: return None
    if isinstance(v, datetime.datetime): return None
    try:
        f = float(v)
        return int(f) if typ == "int" else round(f, 4)
    except: return None

def _date(v):
    if isinstance(v, datetime.datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, str) and "-" in v:  return v[:10]
    return None

def _read_cell(ws, cell_addr, typ="float"):
    v = ws[cell_addr].value
    if   typ == "date":  return _date(v)
    elif typ == "int":   return _num(v, "int")
    elif typ == "float": return _num(v, "float")
    elif typ == "str":   return str(v).strip() if v else None
    return v

# ── 版型版本偵測（執行 layout_detection 規則）──────────────────
def _detect_overrides(ws, mapping: dict) -> dict:
    overrides = {}
    for rule in mapping.get("layout_detection", []):
        cond = rule.get("condition", {})
        cell_val = str(ws[cond["cell"]].value or "")
        if "contains" in cond and cond["contains"] in cell_val:
            overrides.update(rule.get("overrides", {}))
    return overrides

# ── 解析單一週表 ──────────────────────────────────────────────
def _parse_weekly_sheet(ws, mapping: dict) -> dict:
    wm      = mapping["weekly"]
    ovr     = _detect_overrides(ws, mapping)
    # 合併 overrides：設定檔欄位 + 本週版型調整
    eff = {k: ovr.get(k, v) for k, v in wm.items()}

    def get(key):
        cfg = eff.get(key)
        if cfg is None: return None
        return _read_cell(ws, cfg["cell"], cfg.get("type", "float"))

    # remain_parking：新版型 S31 有「未售」→ AC31=合計，舊版 → X31
    if "remain_parking" in eff:
        s31 = str(ws["S31"].value or "")
        if "未售" in s31:
            remain_parking = _read_cell(ws, "AC31", "int")   # 新版：車位合計
        else:
            remain_parking = _read_cell(ws, "X31", "int")    # 舊版
    else:
        remain_parking = None

    return {
        "week_start":          get("week_start"),
        "week_end":            get("week_end"),
        "week_no":             get("week_no"),
        "visitors_week":       get("visitors_week"),
        "visitors_cum":        get("visitors_cum"),
        "calls_week":          get("calls_week"),
        "ordered_week_units":  get("ordered_week_units"),
        "deposit_week_units":  get("deposit_week_units"),
        "signed_week_units":   get("signed_week_units"),
        "signed_week_amount":  get("signed_week_amount"),
        "cancelled_week_units":get("cancelled_week_units"),
        "signed_cum_units":    get("signed_cum_units"),
        "signed_cum_amount":   get("signed_cum_amount"),
        "deals_week":          get("deals_week"),
        "remain_units":        get("remain_units"),
        "remain_parking":      remain_parking,
        "budget_media":        get("budget_media"),
        "budget_personnel":    get("budget_personnel"),
        "budget_onsite":       get("budget_onsite"),
        "budget_company":      get("budget_company"),
        "budget_bonus":        get("budget_bonus"),
        "budget_total":        get("budget_total"),
        "exec_rate_budget":    get("exec_rate_budget"),
        "exec_rate_sales":     get("exec_rate_sales"),
        "billable_fee":        get("billable_fee"),
        "billed_fee":          get("billed_fee"),
        "this_month_request":  get("this_month_request"),
    }

# ── 解析銷況表 ────────────────────────────────────────────────
def _parse_unit_status(wb, mapping: dict) -> dict:
    cfg  = mapping.get("unit_status", {})
    sn   = cfg.get("sheet_name", "銷況表")
    if sn not in wb.sheetnames:
        return {}
    ws = wb[sn]

    hdr_row  = cfg.get("header_row", 3)
    fl_col   = cfg.get("floor_col", "A")
    d_start  = cfg.get("data_start_row", 6)
    col_s    = cfg.get("data_start_col", "D")
    col_e    = cfg.get("data_end_col",   "S")
    abs_row  = cfg.get("absorption_row", 36)
    rem_row  = cfg.get("remain_row",     35)

    # 讀戶別 header
    col_start = openpyxl.utils.column_index_from_string(col_s)
    col_end   = openpyxl.utils.column_index_from_string(col_e)
    units = [ws.cell(row=hdr_row, column=c).value for c in range(col_start, col_end+1)]

    # 讀每層每戶狀態
    matrix = []
    for row in ws.iter_rows(min_row=d_start, values_only=False):
        floor = row[0].value
        if not floor or "去化" in str(floor) or "剩餘" in str(floor): break
        fl = str(floor).strip()
        for i, c in enumerate(range(col_start-1, col_end)):
            if i >= len(units): break
            status = str(row[c].value or "").strip() or "未售"
            if units[i]:
                matrix.append({
                    "floor": fl,
                    "unit":  str(units[i]),
                    "status": status,
                })

    # 讀去化率與剩餘統計
    absorption = {}
    remain     = {}
    if abs_row:
        row_abs = ws[abs_row]
        row_rem = ws[rem_row]
        for cell in row_abs:
            v = _num(cell.value)
            if v is not None:
                col_l = openpyxl.utils.get_column_letter(cell.column)
                absorption[col_l] = round(v * 100, 1)
        for cell in row_rem:
            v = _num(cell.value)
            if v is not None:
                col_l = openpyxl.utils.get_column_letter(cell.column)
                remain[col_l] = int(v)

    return {"matrix": matrix, "absorption": absorption, "remain": remain}

# ── 解析銷售人員業績 ──────────────────────────────────────────
def _parse_salespeople(ws, mapping: dict) -> list:
    cfg = mapping.get("salespeople", {})
    if not cfg: return []
    start = cfg.get("data_start_row", 39)
    end   = cfg.get("data_end_row",   46)
    people = []
    for r in range(start, end+1):
        name = ws.cell(row=r, column=1).value
        if not name or "合計" in str(name) or "櫃台" in str(name) or "結案" in str(name):
            continue
        people.append({
            "name":          str(name).strip(),
            "cum_visitors":  _num(ws.cell(row=r, column=openpyxl.utils.column_index_from_string(cfg.get("col_cum_visitors","K"))).value, "int"),
            "cum_deals":     _num(ws.cell(row=r, column=openpyxl.utils.column_index_from_string(cfg.get("col_cum_deals","L"))).value, "int"),
            "conv_rate":     _num(ws.cell(row=r, column=openpyxl.utils.column_index_from_string(cfg.get("col_conv_rate","M"))).value),
        })
    return people

# ── 主入口 ────────────────────────────────────────────────────
def parse_project_workbook(filepath: str, project_name: str = None) -> dict:
    wb   = openpyxl.load_workbook(filepath, data_only=True)
    # 自動偵測對映設定檔
    if not project_name:
        for f in sorted(MAPPING_DIR.glob("*_mapping.yaml")):
            nm = f.stem.replace("_mapping", "")
            if nm in pathlib.Path(filepath).name:
                project_name = nm
                break
    if not project_name:
        return {"status": "error", "message": "無法比對任何對映設定檔，請手動指定 project_name"}

    mapping = load_mapping(project_name)
    fp_cfg  = mapping.get("fingerprint", {})
    pat     = re.compile(fp_cfg.get("sheet_pattern", r"^\d{4}-\d{4}$"))
    week_sheets = sorted([s for s in wb.sheetnames if pat.match(s)])

    # 解析每週
    weekly_data = {}
    for sn in week_sheets:
        ws   = wb[sn]
        rec  = _parse_weekly_sheet(ws, mapping)
        rec["sheet"] = sn
        # 以 week_start 為主鍵，fallback 到 sheet name
        key  = rec.get("week_start") or sn
        sp   = _parse_salespeople(ws, mapping)
        if sp: rec["salespeople"] = sp
        weekly_data[key] = rec

    # 銷況表（最新狀態，只存一份）
    unit_status = _parse_unit_status(wb, mapping)

    # 轉換成月份索引（113/10 → "113.10"）
    def to_month(date_str):
        if not date_str: return None
        try:
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            roc_yr = dt.year - 1911
            return f"{roc_yr}.{dt.month:02d}"
        except: return None

    # 每月彙總：取當月最後一週的 remain/deals
    monthly = {}
    for key, rec in sorted(weekly_data.items()):
        m = to_month(rec.get("week_start"))
        if m: monthly[m] = rec   # 後者覆蓋前者 → 月末最新值

    return {
        "status":      "ok",
        "project_name": project_name,
        "weekly":       weekly_data,
        "monthly":      monthly,
        "unit_status":  unit_status,
        "parsed_weeks": len(weekly_data),
        "sheet_names":  week_sheets,
    }

# ── 合併到 db.json ────────────────────────────────────────────
def merge_to_db(result: dict):
    if result.get("status") != "ok":
        return result

    db = json.loads(DB_JSON.read_text(encoding="utf-8")) if DB_JSON.exists() else {}
    pname = result["project_name"]

    # 建立 projects 區（若無）
    db.setdefault("projects", {})
    db["projects"].setdefault(pname, {})

    proj = db["projects"][pname]
    proj["weekly"]      = result["weekly"]
    proj["unit_status"] = result["unit_status"]

    # 把月彙總同步補進 portfolio（讓儀表板抽屜可顯示週細節）
    proj["monthly"] = result["monthly"]
    proj["updated_at"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    DB_JSON.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"status": "ok", "merged_to_db": True, **result}

# ── CLI ───────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("用法：python parse_project.py <Excel路徑> [建案名稱]")
        sys.exit(1)
    filepath = sys.argv[1]
    pname    = sys.argv[2] if len(sys.argv) > 2 else None
    r = parse_project_workbook(filepath, pname)
    if r["status"] == "ok":
        merged = merge_to_db(r)
        print(f"✅ 解析完成")
        print(f"   建案：{r['project_name']}")
        print(f"   週數：{r['parsed_weeks']}")
        print(f"   銷況表戶數：{len(r['unit_status'].get('matrix',[]))}")
        print(f"   月份：{sorted(r['monthly'].keys())}")
        print(f"   db.json 已更新")
        print()
        print("📌 下一步：")
        print("   git add db.json")
        print(f'   git commit -m "更新 {r["project_name"]} 單案資料"')
        print("   git push")
    else:
        print(f"❌ 錯誤：{r.get('message')}")
        sys.exit(1)
